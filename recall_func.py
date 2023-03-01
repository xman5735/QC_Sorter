import os
import openpyxl

def read_file(lot_num_in, color_in, profile_in, line_num_in):
    

    #find path to users desktop
    #desktop = os.path.join(os.path.expanduser('~'), 'Desktop')

    # Generate folder path
    folder_path_create = r'\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Reports\\' + 'Line ' + line_num_in + '\\' + profile_in.replace("/", "-") + '\\' + color_in + '\\'
    #files_folder = os.path.join(desktop, folder_path_create)
    files_folder = folder_path_create
    file_name = str(lot_num_in) + '.xlsx'
    file_path = os.path.join(files_folder, file_name)
    
    

    # Function to find the column where the data should be read from
    def read_column(sheet):
        # Initialize an empty list to store the column data
        column_data = {}
        ws = sheet
        # Iterate over the columns in the worksheet
        for col in range(2, ws.max_column + 1):
            # Check if the first row of the column is empty
            if not ws.cell(row=1, column=col).value:
                # If it is empty, break out of the loop
                break
            
            # Extract the values from the first 32 rows of the column
            values = []
            for row in range(1, 33):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    values.append(cell_value)
                else:
                    values.append(0)
            
            # Add the column data to the dictionary
            column_data[ws.cell(row=1, column=col).value] = values

        # Print the resulting dictionary
        print(column_data)
        return column_data

    def save_times_as_tuple(column_data):
        keys_tuple = tuple(column_data.keys())
        return keys_tuple
    
    # Check if folder path exists, return false if it doesnt
    if not os.path.exists(file_path):
        file_exists = False
        print(file_path)
    if os.path.exists(file_path):
        file_exists = True
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active
        # Find the column to write the data
        column_data = read_column(sheet)

    print(file_path)
    keys_tuple = save_times_as_tuple(column_data)
    print(keys_tuple)

    return file_exists, column_data, keys_tuple, file_path
    #return column_times, column_data, file_exists, file_path


def pull_from_file(columns, file_path):
    # Loads the workbook from the specified file path as read-only
    wb = openpyxl.load_workbook(file_path, read_only=True)
    
    # Selects the active sheet in the workbook
    sheet = wb.active

    # Initializes an empty dictionary to store column values
    column_values = {}

    # Loops over each column specified in the columns list
    for col_letter, value in columns:
        # Initializes an empty list in the column_values dictionary for the current column
        column_values[col_letter] = []
        
        # Loops over each cell in the current column and appends the cell's value (or the string "null" if the cell is empty) to the corresponding list in the column_values dictionary
        for cell in sheet[col_letter]:
            if cell.value:
                column_values[col_letter].append(cell.value)
            else:
                column_values[col_letter].append("null")

    # Loops over each column in the column_values dictionary and prints the column letter and its corresponding list of values
    for col_letter, values in column_values.items():
        #print(f"Column {col_letter}: {values}")
        
        # Returns the column_values dictionary
        return(column_values)
    
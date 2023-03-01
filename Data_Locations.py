import os
import openpyxl


def averageLab(colorTuple):
    #colorTuple_count = colorTuple.count
    if len(colorTuple) == 9:
        color_L1, color_L2, color_L3, color_a1, color_a2, color_a3, color_b1, color_b2, color_b3 = colorTuple
        avgL = ((float(color_L1) + float(color_L2) + float(color_L3))/3)
        avga = ((float(color_a1) + float(color_a2) + float(color_a3))/3)
        avgb = ((float(color_b1) + float(color_b2) + float(color_b3))/3)
        return avgL, avga, avgb
    if len(colorTuple) == 3:
        avg_L_set, avg_a_set, avg_b_set = colorTuple
        avgL = float(avg_L_set)
        avga = float(avg_a_set)
        avgb = float(avg_b_set)

        return avgL, avga, avgb
    else:
        print("color average error")

    

# This function is checking the color standard range of a given color, specified by the parameter `colorOut`.
# The `color_cell_ranges` dictionary maps colors to ranges of cells in an Excel spreadsheet, which stores the top and bottom
# bounds of the standard range for each color. 
# The function then opens the specified Excel file and reads the values of the cells in the specified range, storing them in 
# variables `result_L_top`, `result_a_top`, `result_b_top`, `result_L_bottom`, `result_a_bottom`, and `result_b_bottom`. 
# Finally, the function returns these values.
def check_color_standard_range(lot_avg_L, lot_avg_a, lot_avg_b, colorOut):
    color_cell_ranges = {
        'White': ('Q5', 'R5', 'S5', 'Q6', 'R6', 'S6'),
        'Yellow': ('Q9', 'R9', 'S9', 'Q10', 'R10', 'S10'),
        'Light Grey': ('Q13', 'R13', 'S13', 'Q14', 'R14', 'S14'),
        'Weathered Wood': ('Q17', 'R17', 'S17', 'Q18', 'R18', 'S18'),
        'Dark Grey': ('Q21', 'R21', 'S21', 'Q22', 'R22', 'S22'),
        'Lime Green': ('Q25', 'R25', 'S25', 'Q26', 'R26', 'S26'),
        'Aruba Blue': ('Q29', 'R29', 'S29', 'Q30', 'R30', 'S30'),
        'Turf Green': ('Q33', 'R33', 'S33', 'Q34', 'R34', 'S34'),
        'Cherry Wood': ('Q37', 'R37', 'S37', 'Q38', 'R38', 'S38'),
        'Cardinal Red': ('Q41', 'R41', 'S41', 'Q42', 'R42', 'S42'),
        'Patriot Blue': ('Q45', 'R45', 'S45', 'Q46', 'R46', 'S46'),
        'Tudor Brown': ('Q49', 'R49', 'S49', 'Q50', 'R50', 'S50'),
        'Black': ('Q53', 'R53', 'S53', 'Q54', 'R54', 'S54')
    }
    color_cell_range = color_cell_ranges[colorOut]

    file_path = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Standards\S-001 Colorimeter Profile Standard.xlsx"
    df = openpyxl.load_workbook(file_path)
    sheet = df.active

    result_L_top, result_a_top, result_b_top, result_L_bottom, result_a_bottom, result_b_bottom = [sheet[cell].value for cell in color_cell_range]

    return result_L_top, result_a_top, result_b_top, result_L_bottom, result_a_bottom, result_b_bottom

def color_range_check(color_result_tuple):
    # Unpack the color_result_tuple into separate variables
    result_L_top, result_a_top, result_b_top, result_L_bottom, result_a_bottom, result_b_bottom, AvgL, Avga, Avgb = color_result_tuple

    def check_range(value, top, bottom):
        # Check if the value is within the range
        # If the value is greater than top, return the difference and False
        if value > top:
            return abs(value - top), False
        # If the value is smaller than bottom, return the difference and False
        elif value < bottom:
            return -abs(value - bottom), False
        # If the value is within the range, return 0 and True
        else:
            return 0, True

    # Check the range of Average L, a and b values
    L_range, pass_rangeL = check_range(AvgL, result_L_top, result_L_bottom)
    a_range, pass_rangea = check_range(Avga, result_a_top, result_a_bottom)
    b_range, pass_rangeb = check_range(Avgb, result_b_top, result_b_bottom)

    # Pack the range and pass values into a tuple
    pass_range_tuple = (L_range, a_range, b_range, pass_rangeL, pass_rangea, pass_rangeb)
    return pass_range_tuple

# This function is checking the profile standard range of a given profile, specified by the parameter `profileOut`.
# The `profile_cell_ranges` dictionary maps profiles to ranges of cells in an Excel spreadsheet, which stores the top and bottom
# bounds of the standard range for each profile. 
# The function then opens the specified Excel file and reads the values of the cells in the specified range, storing them in 
# variables `result_L_top`, `result_a_top`, `result_b_top`, `result_L_bottom`, `result_a_bottom`, and `result_b_bottom`. 
# Finally, the function returns these values.
def check_profile_standard_range(profile_width, profile_edge, profile_middle, profile_med, boardOut):
    profile_cell_ranges = {
    '1/2x8': ('L14', 'K14', 'L9', 'K9'),
    '3/4x1-3/4': ('L15', 'K15', 'L4', 'K4'),
    '3/4x2-5/8': ('L14', 'K14', 'L6', 'K6'),
    '3/4x3-1/2': ('L14', 'K14', 'L7', 'K7'),
    '3/4x5-1/2': ('L14', 'K14', 'L8', 'K8'),
    '1x5-1/2': ('L16', 'K16', 'L8', 'K8'),
    '1-1/8x3-1/2': ('L17', 'K17', 'L7', 'K7'),
    '1-1/2x1-1/2': ('L18', 'K18', 'L3', 'K3'),
    '1-1/2x2-1/2': ('L18', 'K18', 'L5', 'K5'),
    '1-1/2x3-1/2': ('L18', 'K18', 'L7', 'K7'),
    '1-1/2x5-1/2': ('L18', 'K18', 'L8', 'K8'),
    '1-1/2x9-1/2': ('L18', 'K18', 'L10', 'K10'),
    '2-1/2x2-1/2': ('L19', 'K19', 'L5', 'K5'),
    '3-1/2x3-1/2': ('L20', 'K20', 'L7', 'K7'),
    #'Bench Frame': ('L14', 'K14', 'L9', 'K9')
    }
    profile_cell_range = profile_cell_ranges[boardOut]

    file_path = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Standards\S-004 Profile Thickness Δ Standard.xlsx"
    df = openpyxl.load_workbook(file_path)
    sheet = df.active

    result_thickness_T, result_thickness_B, result_width_T, result_width_B = [sheet[cell].value for cell in profile_cell_range]

    return result_thickness_T, result_thickness_B, result_width_T, result_width_B

def profile_range_check(pass_profile_tuple):
    # Unpack the profile_result_tuple into separate variables
    result_thickness_T, result_thickness_B, result_width_T, result_width_B, profile_width, profile_edge, profile_middle, profile_med = pass_profile_tuple

    def check_range(value, top, bottom):
        value = float(value)
        top = float(top)
        bottom = float(bottom)
        # Check if the value is within the range
        # If the value is greater than top, return the difference and False
        if value > top:
            return abs(value - top), False
        # If the value is smaller than bottom, return the difference and False
        elif value < bottom:
            return -abs(value - bottom), False
        # If the value is within the range, return 0 and True
        else:
            return 0, True

    def check_difference(valueOne, valueTwo, difference):
        valueOne = float(valueOne)
        valueTwo = float(valueTwo)
        difference = float(difference)

        if abs(valueOne - valueTwo) > difference:
            return abs(valueOne - valueTwo), False
        else:
            return 0, True

    # Check the range of values against the input numbers
    width_range, pass_range_width = check_range(profile_width, result_width_T, result_width_B)
    edge_range, pass_range_edge = check_range(profile_edge, result_thickness_T, result_thickness_B)
    middle_range, pass_range_middle = check_range(profile_middle, result_thickness_T, result_thickness_B)

    # Check the input numbers against eachother
    med_range, pass_range_med = check_difference(profile_med, profile_edge, 0.02)
    delta_range, pass_range_delta = check_difference(profile_edge, profile_middle, 0.02)

    # Pack the range and pass values into a tuple
    pass_profile_tuple = (width_range, pass_range_width, edge_range, pass_range_edge, middle_range, pass_range_middle, med_range, pass_range_med, delta_range, pass_range_delta )
    return pass_profile_tuple
def format_excel(format_excel_tuple):
    lot_num, color, profile, pallet_num, date_produced, line_num, hour_sampled, date_sampled = format_excel_tuple 

    # Get the data from previous functions
    format_data = ["Sampled Time:", "Sampled Date:", ".", "Lot Num:", "Profile:", "Color:", "Line Num:", "Date Produced:", "Pallet Num:", ".", "Density:", "Profile Width:", "Result Width:", "Profile Edge:", "Result Edge:",
                    "Profile Mid:", "Result Mid:", "Profile Med:", "Result Med:", ".", "Color L:", "Result L:", "Color a:", "Result a:", "Color b:", "Result b:", "Delta:", "Color Lot #:", "Notes", ".", "Image Path:", "ERC Path:"]

    #Create excel file workbook, check if it exists
    
    # Select the first sheet in the workbook
    #ws = wb.active
    # Get the number of rows in the sheet
    #max_row = ws.max_row

    #for i, value in enumerate(format_data):
    #    ws.cell(row=max_row + 1, column = i, value=value)

    def append_to_excel(file_path, data_list):
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)

        # Select the first sheet in the workbook
        sheet = workbook.active

        # set the width of the column
        sheet.column_dimensions['A'].width = 20

        # Append the data to the sheet, one value at a time
        for i, value in enumerate(data_list):
            sheet.cell(row=i + 1, column=1, value=value)

        # Save the changes to the workbook
        workbook.save(file_path)

    
    # Save the Excel file on the user's desktop in a folder called "Files"
    # Desktop path
    #desktop = os.path.join(os.path.expanduser('~'), 'Desktop')

    # Generate folder path
    folder_path_create = r'\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Reports\\' + 'Line ' + line_num + '\\' + profile.replace("/", "-") + '\\' + color + '\\'
    #files_folder = os.path.join(desktop, folder_path_create)
    files_folder = folder_path_create
    # Check if folder path exists, create it if it doesn't
    if not os.path.exists(files_folder):
        os.makedirs(files_folder)

    # Generate file path
    file_path = os.path.join(files_folder, lot_num + '.xlsx')

    # Check if file exists, skip creation if it does
    if not os.path.exists(file_path):
        # Create the file if it doesn't exist
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(file_path)

    wb.save(file_path)

    append_to_excel(file_path, format_data)

    return file_path

def write_to_excel(file_path, to_excel_tuple):
    # Unpack the tuple into individual variables
    (lot_num, color, profile, pallet_num, date_produced, line_num, hour_sampled, date_sampled, density, avg_L, avg_a, avg_b, range_L, 
    range_a, range_b, delta, width_entry, edge_entry, middle_entry, med_entry, widthStr, edgeStr, middleStr, medStr, deltaStr, color_lot, notes, img_file_name, ERC_file_name) = to_excel_tuple

    # Format the data into a list
    format_data = [hour_sampled, date_sampled, ".", lot_num, profile, color, line_num, date_produced, pallet_num, ".", density, width_entry, widthStr, edge_entry, edgeStr,
                    middle_entry, middleStr, med_entry, medStr, ".", avg_L, range_L, avg_a, range_a, avg_b, range_b, deltaStr, color_lot, notes, ".", img_file_name, ERC_file_name]

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Function to append the data to the excel sheet
    def append_to_excel(file_path, data_list, column_set):
            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_path)

            # Select the first sheet in the workbook
            sheet = workbook.active

            # Set the width of column A
            sheet.column_dimensions['A'].width = 20

            # Append the data to the sheet, one value at a time
            for i, value in enumerate(data_list):
                sheet.cell(row=i + 1, column=column_set, value=value)

            # Save the changes to the workbook
            workbook.save(file_path)

    # Function to find the column where the data should be written
    def find_column_time(sheet, target_value, target_date):
        column = 1
        # Loop through each column
        while True:
            cell_value = sheet.cell(row=1, column=column).value
            cell_value_date = sheet.cell(row=2, column=column).value

            # Check if the column already has an entry with the same hour and date
            if cell_value == target_value:
                if cell_value_date == target_date:
                    print('Previous entry found in ' + str(column) + ' and will be updated')
                    return column
            # Check if the column is empty
            elif cell_value is None:
                print('No entry found and will be updated at ' + str(column))
                return column
            column += 1

    # Find the column to write the data
    column = find_column_time(sheet, hour_sampled, date_sampled)

    # Write the data to the excel sheet
    append_to_excel(file_path, format_data, column)

def write_color(to_excel_tuple):
        
    (lot_num, lot_color_var, lot_profile_var, lot_pallet_num_var, lot_date_produced_var, lot_line_num_var, hours, dateStr,
                    densityStr, lot_avg_L, lot_avg_a, lot_avg_b, lot_range_L, lot_range_a, lot_range_b, deltaStr,
                        width_entry, edge_entry, middle_entry, med_entry, widthStr, edgeStr, middleStr, medStr, deltaStr, 
                            color_lotStr, notes_entry, img_file_name,  ERC_file_name) = to_excel_tuple

    def getfilepath(lot_color_var):
        # Save the Excel file on the user's desktop in a folder called "Files"
        # Desktop path
        #desktop = os.path.join(os.path.expanduser('~'), 'Desktop')

        # Generate folder path
        folder_path_create = r'\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\\Logs\\Color_logs\\'
        #files_folder = os.path.join(desktop, folder_path_create)
        files_folder = folder_path_create
        # Check if folder path exists, create it if it doesn't
        if not os.path.exists(files_folder):
            os.makedirs(files_folder)

        # Generate file path
        file_path = os.path.join(files_folder, lot_color_var + '.xlsx')

        # Check if file exists, skip creation if it does
        if not os.path.exists(file_path):
            # Create the file if it doesn't exist
                wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(file_path)

        wb.save(file_path)

        return file_path
    
    def format_file(file_path):
        data_list =  ["Lot Num:", "Color:", "Line Num:", "Date Produced:", "Sampled Time:", "Color L:", "Result L:", "Color a:", "Result a:", "Color b:", "Result b:", "Color Lot #:"]
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)

        # Select the first sheet in the workbook
        sheet = workbook.active

        # set the width of the column
        sheet.column_dimensions['A'].width = 20

        # Append the data to the sheet, one value at a time
        for i, value in enumerate(data_list):
            sheet.cell(row=i + 1, column=1, value=value)

        # Save the changes to the workbook
        workbook.save(file_path)
        return(sheet)
    
    def find_column(sheet, target_value, target_date):
        column = 1
        # Loop through each column
        while True:
            cell_value = sheet.cell(row=1, column=column).value
            cell_value_date = sheet.cell(row=5, column=column).value

            # Check if the column already has an entry with the same hour and date
            if cell_value == target_value:
                if cell_value_date == target_date:
                    print('Previous entry found in ' + str(column) + ' and will be updated')
                    return column
            # Check if the column is empty
            elif cell_value is None:
                print('No entry found and will be updated at ' + str(column))
                return column
            column += 1
    file_path = getfilepath(lot_color_var)
    sheet = format_file(file_path)
    column_set = find_column(sheet, lot_num, hours)
    # Function to append the data to the excel sheet

    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet in the workbook
    sheet = workbook.active

    data_list = lot_num, lot_color_var, lot_line_num_var, dateStr, hours, lot_avg_L, lot_range_L, lot_avg_a, lot_range_a, lot_avg_b, lot_range_b, color_lotStr
    # Append the data to the sheet, one value at a time
    for i, value in enumerate(data_list):
        sheet.cell(row=i + 1, column=column_set, value=value)

    # Save the changes to the workbook
    workbook.save(file_path)

def write_profile(to_excel_tuple):
            
    (lot_num, lot_color_var, lot_profile_var, lot_pallet_num_var, lot_date_produced_var, lot_line_num_var, hours, dateStr,
                    densityStr, lot_avg_L, lot_avg_a, lot_avg_b, lot_range_L, lot_range_a, lot_range_b, deltaStr,
                        width_entry, edge_entry, middle_entry, med_entry, widthStr, edgeStr, middleStr, medStr, deltaStr, 
                            color_lotStr, notes_entry, img_file_name,  ERC_file_name) = to_excel_tuple

    def getfilepath(lot_profile_var):
        # Save the Excel file on the user's desktop in a folder called "Files"
        # Desktop path
        #desktop = os.path.join(os.path.expanduser('~'), 'Desktop')

        # Generate folder path
        folder_path_create = r'\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Logs\\Profile_logs\\'
        #files_folder = os.path.join(desktop, folder_path_create)
        files_folder = folder_path_create
        # Check if folder path exists, create it if it doesn't
        if not os.path.exists(files_folder):
            os.makedirs(files_folder)

        # Generate file path
        file_path = os.path.join(files_folder, lot_profile_var.replace("/", "-") + '.xlsx')

        # Check if file exists, skip creation if it does
        if not os.path.exists(file_path):
            # Create the file if it doesn't exist
                wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(file_path)

        wb.save(file_path)

        return file_path
    
    def format_file(file_path):
        data_list =  ["Lot Num:", "Profile:", "Line Num:", "Date Produced:", "Sampled Time:", "Width", "Edge", "Middle", "Med", "Delta"]
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)

        # Select the first sheet in the workbook
        sheet = workbook.active

        # set the width of the column
        sheet.column_dimensions['A'].width = 20

        # Append the data to the sheet, one value at a time
        for i, value in enumerate(data_list):
            sheet.cell(row=i + 1, column=1, value=value)

        # Save the changes to the workbook
        workbook.save(file_path)
        return(sheet)
    
    def find_column(sheet, target_value, target_date):
        column = 1
        # Loop through each column
        while True:
            cell_value = sheet.cell(row=1, column=column).value
            cell_value_date = sheet.cell(row=5, column=column).value

            # Check if the column already has an entry with the same hour and date
            if cell_value == target_value:
                if cell_value_date == target_date:
                    print('Previous entry found in ' + str(column) + ' and will be updated')
                    return column
            # Check if the column is empty
            elif cell_value is None:
                print('No entry found and will be updated at ' + str(column))
                return column
            column += 1
    file_path = getfilepath(lot_profile_var)
    sheet = format_file(file_path)
    column_set = find_column(sheet, lot_num, hours)
    # Function to append the data to the excel sheet

    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet in the workbook
    sheet = workbook.active

    data_list = lot_num, lot_profile_var, lot_line_num_var, dateStr, hours, width_entry, edge_entry, middle_entry, med_entry, deltaStr
    # Append the data to the sheet, one value at a time
    for i, value in enumerate(data_list):
        sheet.cell(row=i + 1, column=column_set, value=value)

    # Save the changes to the workbook
    workbook.save(file_path)

def set_photo_place():
    # Save the Excel file on the user's desktop in a folder called "Files"
    # Desktop path
    #desktop = os.path.join(os.path.expanduser('~'), 'Desktop')

    # Generate folder path
    folder_path_create = r'\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Photos\\Pallets\\'
    #files_folder = os.path.join(desktop, folder_path_create)
    files_folder = folder_path_create
    # Check if folder path exists, create it if it doesn't
    if not os.path.exists(files_folder):
        os.makedirs(files_folder)

    return files_folder
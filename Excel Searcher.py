import os
import glob
import openpyxl
import shutil

# Set the path to the directory containing the Excel files
dir_path = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Reports\\"

# Set the path to the user's desktop
desktop_path = os.path.expanduser("~") + "\\Desktop\\"

# Define a function to recursively search for Excel files
def find_excel_files(dir_path):
    excel_files = []
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            if file.endswith(".xlsx"):
                excel_files.append(os.path.join(root, file))
    return excel_files

# Find all Excel files in the directory and its subdirectories
excel_files = find_excel_files(dir_path)

# Loop through each Excel file and read the second row of each column
for file in excel_files:
    wb = openpyxl.load_workbook(file)
    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=2, column=col)
            text = cell.value
            if text is not None:
                # Write the text to a file on the user's desktop
                with open(desktop_path + "excel_text.txt", "a") as f:
                    f.write(text + "\n")

# Move the file from the desktop to the same directory as the Excel files
#shutil.move(desktop_path + "excel_text.txt", dir_path)

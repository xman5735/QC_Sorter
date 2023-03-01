import os
import datetime
from pathlib import Path
import xlsxwriter
import openpyxl

# Characters used for encode and decode. Modifying will result in lot number generation changes and failure to decode past lot numbers.
charset = 'ABCDEFGHJKLMNPQRSTVWXYZ23456789#!'

def getColorDecode(color):
    # Dictionary of color codes and their corresponding names
    colors = {'00001': 'White', '00010': 'Yellow', '00011': 'Light Grey', 
              '00100': 'Weathered Wood', '00101': 'Dark Grey', '00110': 'Lime Green', 
              '00111': 'Aruba Blue', '01000': 'Turf Green', '01001': 'Cherry Wood', 
              '01010': 'Cardinal Red', '01011': 'Patriot Blue', '01100': 'Tudor Brown', 
              '01101': 'Black'}
    
    # Get the color name from the dictionary based on the input color code
    color_num = colors.get(color)
    
    # Check if the input color code exists in the dictionary
    if color_num:
        # If exists, print and return the color name
        print('The board color is ', color_num)
        return color_num
    else:
        # If not, print an error message
        print("Number not listed as a color, check lot number")

def getBoardSizeDecode(boardSize):
    # A dictionary that maps binary encodings to board sizes
    board_sizes = {
        '00001': '1/2x8',
        '00010': '3/4x1-3/4',
        '00011': '3/4x2-5/8',
        '00100': '3/4x3-1/2',
        '00101': '3/4x5-1/2',
        '00110': '1x5-1/2',
        '00111': '1-1/8x3-1/2',
        '01000': '1-1/2x1-1/2',
        '01001': '1-1/2x2-1/2',
        '01010': '1-1/2x3-1/2',
        '01011': '1-1/2x5-1/2',
        '01100': '1-1/2x9-1/2',
        '01101': '2-1/2x2-1/2',
        '01110': '3-1/2x3-1/2',
        '01111': 'Bench Frame',
    }
    
    # Look up the board size in the dictionary using the given binary encoding
    boardNum = board_sizes.get(boardSize, None)
    
    # If the board size is not found, print an error message and return None
    if boardNum is None:
        print("The board number is incorrect, check lot number")
    # If the board size is found, print the size and return it
    else:
        print('The board size is ', boardNum)
        return boardNum

#if length of lot number changes (excluding -x) then update here
def getLotNum(lotNumBin):
    # get the last digit of the binary string as the pallet number
    palletNum = lotNumBin[-1:]

    # remove the last 2 characters (the pallet number and the space) from the binary string
    lotNumBin = lotNumBin[:-2]

    # check if the length of the modified binary string is equal to 5
    if len(lotNumBin) == 5:
        # return the binary string and the pallet number if the length is correct
        return lotNumBin, palletNum
    else:
        # print an error message if the length of the binary string is incorrect
        print("Lot Number Incorrect")

#if decimal amount changes, update here (and then the affected 'get')
def getSplit(lotNum):

    # Check if the length of the lot number is 25
    if len(lotNum) == 25:
        # Store the lot number in a string variable
        string = lotNum
        
        # Get the first 5 characters of the string, representing the color code
        color = string[0:5]
        
        # Remove the color code from the string
        colorSize = string[5:]
        
        # Get the first 5 characters of the remaining string, representing the size code
        size = colorSize[0:5]
        
        # Remove the size code from the string
        stringSize = colorSize[5:]
        
        # Get the first 4 characters of the remaining string, representing the line number
        line = stringSize[0:4]
        
        # Remove the line number from the string
        lineSize = stringSize[4:]
        
        # Get the first 11 characters of the remaining string, representing the date
        date = lineSize[0:11]
        
        # Remove the date from the string
        dateSize = lineSize[11:]

        # Return the extracted values (color, size, line, date)
        return color, size, line, date
        
    # If the length of the lot number is not 25, print an error message
    else:
        print("Lot number incorrect, please try again")

#uncomment ## lines if binary string grows. it is maxed. Lot will increase by 3 digits
def decode(ascii_string):
    # Convert the input string to all uppercase characters
    ascii_string = ascii_string.upper()

    # Find the decimal representation of each character in the string
    # using the index of the character in the charset
    decimals = [charset.index(char) for char in ascii_string]
    # Take last decimal which is the final chunk length, and the second to last
    # decimal which is the final chunk, and keep them for later to be padded
    # appropriately and appended.
    ##last_chunk_length, last_decimal = decimals.pop(-1), decimals.pop(-1)
    # Convert each decimal to its binary representation, padded to 6 digits
    bin_string = ''.join([bin(decimal)[2:].zfill(5) for decimal in decimals])

    # Add the last decimal converted to binary padded to the appropriate length
    ##bin_string += bin(last_decimal)[2:].zfill(last_chunk_length)
    #print(bin_string)
    return bin_string

# Function to decode the date from a binary string
def getDate(dateNum):
    # Convert the binary string to a decimal string
    dateNum = int(dateNum, 2)
    dateNum = str(dateNum)

    # Add a leading 0 if the length of the decimal string is not 4
    if len(dateNum) != 4:
        dateNum = '0' + dateNum

    # Extract the first 2 characters of the decimal string as the month
    monthNum = dateNum[0:2]

    # Extract the last 2 characters of the decimal string as the day
    dayNum = dateNum[-2:]
        
    # Return the day and month
    return dayNum, monthNum

# Function to assign a number to a color input
def assignColor(colorInput):
    # Print a message indicating that the function has been entered
    print("Into assignColor")

    # Dictionary to map colors to numbers
    color_to_number = {
        'White': 1,
        'Yellow': 2,
        'Light_Grey': 3,
        'Weathered_Wood': 4,
        'Dark_Grey': 5,
        'Lime_Green': 6,
        'Aruba_Blue': 7,
        'Turf_Green': 8,
        'Cherry_Wood': 9,
        'Cardinal_Red': 10,
        'Patriot_Blue': 11,
        'Tudor_Brown': 12,
        'Black': 13
    }

    # Get the assigned number using the color input as a key
    assignedColor = color_to_number[colorInput]

    # Return the assigned number
    return assignedColor

#can handle up to 31 colors before another binary decimal is needed. 
#Turn numerical number into assigned binary
def getColor(color):
    # Mapping of color number to color name and binary representation
    color_map = {
        '1': ('White', bin(1)),
        '2': ('Yellow', bin(2)),
        '3': ('Light Grey', bin(3)),
        '4': ('Weathered Wood', bin(4)),
        '5': ('Dark Grey', bin(5)),
        '6': ('Lime Green', bin(6)),
        '7': ('Aruba Blue', bin(7)),
        '8': ('Turf Green', bin(8)),
        '9': ('Cherry Wood', bin(9)),
        '10': ('Cardinal Red', bin(10)),
        '11': ('Patriot Blue', bin(11)),
        '12': ('Tudor Brown', bin(12)),
        '13': ('Black', bin(13)),
    }

    # Convert color number to string
    color_str = str(color)

    # Return None, None if color number not found in color_map
    if color_str not in color_map:
        return None, None

    # Get color name and binary representation from color_map
    color_name, color_num = color_map[color_str]

    # Format binary representation to be 5 characters long with leading zeros if necessary
    color_num = str(color_num)[2:].zfill(5)

    # Print color name for reference
    print('The board color is set to ', color_name)

    # Return binary representation and color name
    return color_num, color_name

# This function assigns a numerical value to a profile input based on the mapping in the profile_to_number dictionary.
def assignProfile(profileInput):
    # A dictionary mapping profile strings to their respective numerical values
    profile_to_number = {
    '1/2 x 8': 1,
    '3/4 x 1-3/4': 2,
    '3/4 x 2-5/8': 3,
    '3/4 x 3-1/2': 4,
    '3/4 x 5-1/2': 5,
    '1 x 5-1/2': 6,
    '1-1/8 x 3-1/2': 7,
    '1-1/2 x 1-1/2': 8,
    '1-1/2 x 2-1/2': 9,
    '1-1/2 x 3-1/2': 10,
    '1-1/2 x 5-1/2': 11,
    '1-1/2 x 9-1/2': 12,
    '2-1/2 x 2-1/2': 13,
    '3-1/2 x 3-1/2': 14,
    'Bench Frame': 15
    }
    # Look up the profile input in the dictionary and assign its value to assignedProfile
    assignedProfile = profile_to_number[profileInput]
    # Return the assigned profile value
    return assignedProfile

 #Can handle 31 board sizes until update needed again 
# This function takes a numerical assignedProfile as input and returns the corresponding board size string.
def getBoardSize(assignedProfile):
    # A dictionary mapping numerical assignedProfile values to their corresponding board sizes
    board_sizes = {
        '1': '½ x 8',
        '2': '¾ x 1¾',
        '3': '¾ x 2 5-8',
        '4': '¾ x 3½',
        '5': '¾ x 5½',
        '6': '1 x 5½',
        '7': '1 1-8 x 3½', 
        '8': '1½ x 1½',
        '9': '1½ x 2½',
        '10': '1½ x 3½',
        '11': '1½ x 5½',
        '12': '1½ x 9½',
        '13': '2½ x 2½',
        '14': '3½ x 3½',
        '15': 'Bench Frame',
    }

    # Look up the assignedProfile in the dictionary and assign its value to board_size. 
    # If it is not found, assign "Entered board selection does not match any produced boards".
    board_size = board_sizes.get(str(assignedProfile), "Entered board selection does not match any produced boards")
    boardBin = bin(assignedProfile)
    boardBin = str(boardBin)
    #remove 0b from front of string. result of binary conversion
    boardBin = boardBin[2:]
    boardNum = boardBin.zfill(5)

    # Check if the assignedProfile was not found in the dictionary. If so, print an error message.
    if board_size == "Entered board selection does not match any produced boards":
        print("Entered board selection does not match any produced boards")
        print(board_size)
        return None
    else:
        # Print the corresponding board size and return it.
        print("The selected board size is", board_size)
        return boardNum, board_size

#After 15, will need to increase zfill and decimal size.
# This function takes a line number as input and returns its binary representation.
def getLineNumber(line_number):
    # Convert the input line number to an integer
    lineNumInt  = int(line_number)
    # Print the input line number
    print("The board was extruded on line ", lineNumInt)
    # Convert the integer line number to binary and remove the '0b' prefix
    lineNum = bin(lineNumInt)[2:].zfill(4)
    # Return the binary representation of the line number
    return lineNum

#wont need update until 2100
def getProdDate(dateString):
    proDate = dateString
    month = proDate[0:2]
    proDateShred = proDate[2:]
    day = proDateShred[0:2]
    proDateShred = proDateShred[2:]
    year = proDateShred[0:2]
    #return production date as int and verify
    print("The board was produced on ", proDate)
    proDate = month + day
    proDateInt = int(proDate)
    #remove the 0x and make sure the string is 17 chaRACTERS
    #splitting into seperate sections allows a 1 decimal saving (16) so leaving it.
    proDate = bin(proDateInt)[2:].zfill(11)
    return proDate, day, month, year

#if more than 9 pallets per line are made in a day, needs to update length. This is unlikely
def getPalletNum(pallet_number, num_of_pallets):
    # Assign the value of the first parameter to a new variable "palletNum"
    palletNum = pallet_number
    
    # Assign the value of the second parameter to a new variable "howMany"
    howMany = num_of_pallets
    
    # Print a string with the value of "palletNum"
    print("The pallet number is", palletNum)
    
    # Return a tuple with both "palletNum" and "howMany"
    return(palletNum, howMany)

#functions to modify and create new data. Appends all data to excel file of the month
def printToFile(color, size, line, prodDate, pallet, lot):
    # Convert binary strings to integers
    line = int(line, 2)
    prodDate = int(prodDate, 2)

    # Get today's date and format it fr month and year
    today = datetime.date.today()
    time_string = today.strftime("%m_%y")
    prodDateStr = str(prodDate)
    date_append = str(prodDateStr[:-2]) + "/" + str(prodDateStr[-2:])
    #make inputs into tuplle for enumeration
    outline = ("Lot:", "Date:", "Profile:", "Color:", "Line #:", "Pallet #")
    information = (lot, date_append, size, color, line, pallet)
    # Generate the file name using the date
    file_name = "lot_numbers_" + time_string + ".xlsx"

    # Define the directory where the file will be stored
    #directory = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop\\Files\\Labels\\Records\\')
    directory = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Labels\Records\\"
    # Create the directory if it does not already exist
    if not os.path.exists(directory):
        os.makedirs(directory)

    file_name = directory + file_name

    if os.path.isfile(file_name):

        # Open the Excel file
        workbook = openpyxl.load_workbook(file_name)

        # Select the active worksheet
        worksheet = workbook.active

        # Get the next available column
        next_col = worksheet.max_column + 1

        # Iterate over the tuple and insert each value into a cell in the next available column
        for i, value in enumerate(information):
            cell = worksheet.cell(row=i+1, column=next_col)
            cell.value = value

        workbook.save(file_name)
        print("file appended")

    else:
            # If the file does not exist, create it and add the tuple of values to the first row
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for i, value in enumerate(outline):
            cell = worksheet.cell(row=i+1, column=1)
            cell.value = value
        for i, value in enumerate(information):
            cell = worksheet.cell(row=i+1, column=2)
            cell.value = value
        workbook.save(file_name)
        print("New file created:", file_name)

    return pallet, prodDate


#5 = 31 digits. 6 = 63 digits for charset 
#if the amount of decimals increases, uncomment ## items
def encode(bin_string):
    # Split the string of 1s and 0s into lengths of 5.
    chunks = [bin_string[i:i+5] for i in range(0, len(bin_string), 5)]
    # Store the length of the last chunk so that we can add that as the last bit
    # of data so that we know how much to pad the last chunk when decoding.
    ##last_chunk_length = len(chunks[-1])
    # Convert each chunk from binary into a decimal
    decimals = [int(chunk, 2) for chunk in chunks]
    # Add the length of our last chunk to our list of decimals.
    ##decimals.append(last_chunk_length)
    # Produce an ascii string by using each decimal as an index of our charset.
    ascii_string = ''.join([charset[i] for i in decimals])

    return ascii_string 

def printOut(colorString, palletNum, boardString, encodedLot, makeDate):
    # Convert the color string to uppercase
    color_string = colorString.upper()
    
    #point directory for file storage
    #directory = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop\\Files\\Labels') 
    directory = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Labels\\"
    #if file directory does not exist, create it

    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Directory {directory} created.")
    else:
        print(f"Directory {directory} already exists.")

    # Create the workbook name using the input parameters
    workbook_name = "{}_{}_{}.xlsx".format(color_string, boardString, palletNum)
    workbook_path = directory + "\\" + workbook_name
    #print(workbook_path)
    
    # Create the workbook
    workbook = xlsxwriter.Workbook(workbook_path)
    
    # Add a worksheet to the workbook
    worksheet = workbook.add_worksheet()
    
    # Set the margins for the worksheet
    worksheet.set_margins(left=0.25, right=0.25)
    
    # Define the font sizes for different colors
    font_sizes_color = {
        "WEATHERED WOOD": 60,
        "TUDOR BROWN": 75,
        "CHERRY WOOD": 80,
        "PATRIOT BLUE": 80,
        'WHITE': 90,
        'YELLOW': 90,
        'LIGHT GREY': 90,
        'DARK GREY': 90,
        'LIME GREEN': 90,
        'ARUBA BLUE': 90,
        'TURF GREEN': 85,
        'CARDINAL RED': 80,
        'BLACK': 90
    }
    
    # Define the font sizes for different dimensions
    font_sizes_dimensions = {
        '½ x 8': 90,
        '¾ x 1¾': 90,
        '¾ x 2 5-8': 80,
        '¾ x 3½': 90,
        '¾ x 5½': 90,
        '1 x 5½': 90,
        "1 1-8 x 3½": 70,
        '1½ x 1½': 80,
        '1½ x 2½': 90,
        '1½ x 3½': 90,
        '1½ x 5½': 90,
        '1½ x 9½': 90,
        '2½ x 2½': 90,
        '3½ x 3½': 90,
        'Bench Frame': 60,
    }
    
    
    # Create format for the dimension text
    size1 = workbook.add_format()
    size1.set_font_size(font_sizes_dimensions.get(boardString))
    size1.set_align('center')
    size1.set_align('vcenter')
    
    # Create format for the color text
    size2 = workbook.add_format()
    size2.set_font_size(font_sizes_color.get(color_string))
    size2.set_align('center')
    size2.set_bold()
    
    # Create format for the "Date" and "Lot" text
    size3 = workbook.add_format()
    size3.set_font_size(36)
    size3.set_underline(1)
    size3.set_align('right')
    size3.set_bold()
    
    # Create format *Unused*
    size4 = workbook.add_format()
    size4.set_font_size(36)
    size4.set_underline(1)
    size4.set_align('left')
    size4.set_bold()
    
    # Define the row heights for the worksheet
    row_heights = [105, 140, 85, 40, 105, 140, 85]
    
    # Define the column widths for the worksheet
    column_widths = [42, 10, 42]
    
    # Set the row heights for the worksheet
    for row, height in enumerate(row_heights):
        worksheet.set_row(row, height)
    
    # Set the column widths for the worksheet
    for col, width in enumerate(column_widths):
        worksheet.set_column(col, col, width)
    
    # Prepare the "____ - <boardString>" text
    dims = "____ - {}".format(boardString)
    
    # Store the color text
    color = color_string

    # Prepare the date text
    date = "Date: {}".format(makeDate)

    # Prepare the lot text
    lot = "Lot #: {}".format(encodedLot)

    # Write data to specified excel cell with prepared formating
    worksheet.write('B1', dims, size1)
    worksheet.write('B2', color, size2)
    worksheet.write('A3', date, size3)
    worksheet.write('C3', lot, size3)
    worksheet.write('B5', dims, size1)
    worksheet.write('B6', color, size2)
    worksheet.write('A7', date, size3)
    worksheet.write('C7', lot, size3)

    # Save and close workbook
    workbook.close()
   
#used to increment date by 1 while chacking for a weekend
#comment out noted section if friday has production
def dateInc(day, month, year):

    #define the date into datetime format
    yearS = int(year) + 2000
    monthS = int(month)
    dayS = int(day)
    dateL = datetime.date(yearS, monthS, dayS)

    # Add the number of days to skip weekends based on the current weekday
    # M, T, W, R, F, S, Su
    skip_days = [0, 0, 0, 0, 2, 1, 0][dateL.weekday()]
    dateL += datetime.timedelta(days=skip_days + 1)

    #convert back into string
    day = dateL.strftime("%d")
    month = dateL.strftime("%m")
    year = dateL.strftime("%y")
    print(month, day, year)

    return day, month, year

#used to update the date without going through getProdDate() again
def setDate(day, month, year):
    makeDate = month + day
    makeDateInt = int(makeDate)
    makeDateBin = bin(makeDateInt)[2:].zfill(11)
    #returned makeDate to string for printOut() with correct formating. Since is also updated with the loop.
    makeDate = month + "/" + day + "/" + year
    return makeDate, makeDateBin
import tkinter as tk
import openpyxl
from tkinter import messagebox
from tkcalendar import DateEntry
import openpyxl.utils
import os.path


# Check if the Excel file exists, and open it if it does
file_location = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Logs\Color Masterbatch Logs.xlsx"
if os.path.isfile(file_location):
    wb = openpyxl.load_workbook(file_location)
    ws = wb.active
else:
    # Create a new Excel workbook if the file doesn't exist
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Color Masterbatch Logs"

    # Add headers to the worksheet
    ws.cell(row=1, column=1).value = "Color"
    ws.cell(row=2, column=1).value = "Lot Number"
    ws.cell(row=3, column=1).value = "Date Received"


# Define a function to handle the submit button
def submit():
    # Get the lot number, color, and date received values from the user
    lot_number = str(lot_number_entry.get())
    lot_number = lot_number.upper()
    color = color_var.get()
    date_received = date_received_entry.get()

    # Find the next available column
    column = 1
    while ws.cell(row=1, column=column).value is not None:
        column += 1
        
    # Add the values to the worksheet
    column = ws.max_column + 1
    ws.cell(row=1, column=column).value = color
    ws.cell(row=2, column=column).value = lot_number
    ws.cell(row=3, column=column).value = date_received

    # Set the column width to 110 pixels
    ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 15

    # Save the workbook
    file_location = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Logs\Color Masterbatch Logs.xlsx"
    wb.save(file_location)

    # Show a message box to confirm the submission
    messagebox.showinfo("Success", "Data submitted successfully.")

# Create the main window
root = tk.Tk()
root.title("Color Masterbatch Logs")

# Create a label and entry for the lot number
lot_number_label = tk.Label(root, text="Lot Number:")
lot_number_label.grid(row=0, column=0, padx=10, pady=10)
lot_number_entry = tk.Entry(root)
lot_number_entry.grid(row=0, column=1, padx=10, pady=10)

# Create a label and combobox for the color
color_label = tk.Label(root, text="Color:")
color_label.grid(row=1, column=0, padx=10, pady=10)
color_var = tk.StringVar(root)
color_choices = ['White', 'Yellow', 'Light Grey', 'Weathered Wood', 'Dark Grey', 'Lime Green', 'Aruba Blue', 'Turf Green', 'Cherry Wood', 'Cardinal Red', 'Patriot Blue', 'Tudor Brown', 'Black']
color_combobox = tk.ttk.Combobox(root, textvariable=color_var, values=color_choices)
color_combobox.grid(row=1, column=1, padx=10, pady=10)

# Create a label and calendar for the date received
date_received_label = tk.Label(root, text="Date Received:")
date_received_label.grid(row=2, column=0, padx=10, pady=10)
date_received_entry = DateEntry(root, date_pattern="yyyy-mm-dd")
date_received_entry.grid(row=2, column=1, padx=10, pady=10)

# Create a submit button
submit_button = tk.Button(root, text="Submit", command=submit)
submit_button.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

# Start the main event loop
root.mainloop()

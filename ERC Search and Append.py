import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from tkcalendar import DateEntry
import os
import shutil
import datetime
import glob
import openpyxl

# Function to handle the file selection
def select_file():
    global pdf_file_path
    pdf_file_path = filedialog.askopenfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
    file_label.config(text="Selected File: " + os.path.basename(pdf_file_path))

# Function to handle the submission
def submit():
    start_date = start_date_entry.get_date().strftime("%Y-%m-%d")
    end_date = end_date_entry.get_date().strftime("%Y-%m-%d")
    new_file_name = "ERC_" + start_date + "-" + end_date + ".pdf"
    new_file_path = os.path.join(r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Logs\ERC_PDF", new_file_name)
    print(new_file_path)
    shutil.copy(pdf_file_path, new_file_path)
    update_excel_files(start_date, end_date, new_file_path)

# Function to update Excel files
def update_excel_files(start_date, end_date, new_file_path):
    search_folder = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Reports\\"
    column_count = 0
    for file_path in glob.glob(search_folder + "**/*.xlsx", recursive=True):
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    # Try to convert cell value to datetime object
                    if cell.value:
                        try:
                            date_obj = datetime.datetime.strptime(cell.value, "%Y-%m-%d").date()
                            date_value = date_obj
                            print(date_obj)
                        except ValueError as e:
                            date_value = None
                            print(f"Error: {e}")
                    else:
                        date_value = None
                    # Check if date_value is within the specified range
                    if date_value and start_date <= str(date_value) <= end_date:
                    
                        # Update the cell in row 32 of the same column
                        sheet.cell(row=32, column=cell.column).value = new_file_path
                        wb.save(file_path)
                        column_count += 1
                        break
    # Show message box
    message = f"{column_count} columns have been updated."
    messagebox.showinfo("Success", message)

# Create the GUI
root = tk.Tk()
root.title("PDF and Excel File Manager")

# PDF File Selection
pdf_label = tk.Label(root, text="Select a PDF File:")
pdf_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)

pdf_button = tk.Button(root, text="Browse", command=select_file)
pdf_button.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)

file_label = tk.Label(root, text="Selected File: None")
file_label.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky=tk.W)

# Date Entries
start_date_label = tk.Label(root, text="Start Date:")
start_date_label.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)

start_date_entry = DateEntry(root, date_pattern="yyyy-mm-dd")
start_date_entry.grid(row=2, column=1, padx=5, pady=5)

end_date_label = tk.Label(root, text="End Date:")
end_date_label.grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)

end_date_entry = DateEntry(root, date_pattern="yyyy-mm-dd")
end_date_entry.grid(row=3, column=1, padx=5, pady=5)

# Submit Button
submit_button = tk.Button(root, text="Submit", command=submit)
submit_button.grid(row=4, column=0, padx=5, pady=5)

root.mainloop()

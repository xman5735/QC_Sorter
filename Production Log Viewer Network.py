import tkinter as tk
from tkinter import ttk
import datetime
from tkinter import messagebox
from tkcalendar import DateEntry
from tkinter import filedialog
from PIL import Image
import lot_number_functions
import Data_Locations
import subprocess
import recall_func
import shutil
import sys
import os
import openpyxl


recall_enabled = None

#for setting global varaible recall_enabled. 
def set_recall_enabled(value):
    global recall_enabled
    recall_enabled = value

root = tk.Tk()
root.title("Data Entry")

def last_month_lot():
    today = datetime.date.today()
    time_string = today.strftime("%m_%y")
    #directory = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop\\Files\\Labels\\Records\\')
    directory = r'\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Labels\Records\\'
    file_path = directory + "lot_numbers_" + time_string + ".xlsx"
    workbook = openpyxl.load_workbook(file_path)
    # Select the first sheet
    sheet = workbook.active
    # Create an empty tuple to store the first row of values
    values = ()
    # Loop through each cell in the first row (excluding the first column)
    for cell in sheet.iter_cols(min_row=1, min_col=2, max_col=sheet.max_column):
        # Append the value of each cell to the tuple
        values += (cell[0].value,)
    # Close the Excel file
    workbook.close()
    lot_combo_box['values'] = values
    print("last lot updated")

#run right after start to populate dropdown
root.after(0, last_month_lot)

#update lot number entry with combobox if selected
def update_lot_num_entry(event):
        lot_num.set(lot_combo_box.get())

#Create label to instruct user to enter lot number below 
made_lot_label = tk.Label(root, text="This Months Lots #'s")
made_lot_label.grid(row=0, column=0)


#Create label to instruct user to enter lot number below 
enter_lot_num_label = tk.Label(root, text="Enter Lot #")
enter_lot_num_label.grid(row=0, column=1)
enter_lot_num_label.config(font=("Helvetica", 12))

# text input for "enter lot #"
lot_num = tk.StringVar()
lot_num_entry = tk.Entry(root, textvariable=lot_num)
lot_num_entry.grid(row=1, column=1)
lot_num_entry.config(font=("Helvetica", 12))

#create combobox populated with the last x amount of lot numbers created
last_x_lot = ()
lot_combo_box = ttk.Combobox(root, values=last_x_lot)
lot_combo_box.bind("<<ComboboxSelected>>", update_lot_num_entry)
lot_combo_box.grid(row=1, column=0)
lot_combo_box.config(font=("Helvetica", 12))

# Define the function "color_lot_get"
def color_lot_get():
    # Load the Excel file
    # Define the path to the Excel file
    excel_path = r"\\lcc-fsqb-01.lcc.local\Shares\Green Fox\QC\Logs\Color Masterbatch Logs.xlsx"

    wb = openpyxl.load_workbook(excel_path)

    # Select the active worksheet
    ws = wb.active

    # Get the selected color from the combobox
    selected_color = lot_color_var.get()

    # Find all cells in row 1 that match the selected color
    matching_columns = [col_idx+1 for col_idx in range(ws.max_column) if ws.cell(row=1, column=col_idx+1).value == selected_color]

    # Get the values in row 2 of the matching columns
    matching_values = [ws.cell(row=2, column=col_idx).value for col_idx in matching_columns]

    # Set the values in the combobox
    color_lot_number_entry['values'] = matching_values

# button next to "enter lot #" input
def submit_lot_num():
    set_recall_enabled(False)
        # Get the text entered in the text input field
    text = lot_num_entry.get()
    
    # Set the text of the output label to the entered text
    text = text.upper()
    text = text.replace(" ", "")
    lot_num.set(text)
    lot_num_conf.config(text=text)

    lotNumBin, palletEnd = lot_number_functions.getLotNum(text)
    #convert back to binary
    lotNum = lot_number_functions.decode(lotNumBin)
    #print(lotNum)
    #shred binary into individual parts
    splits = lot_number_functions.getSplit(lotNum)
    #break tuple 'splits' into integers
    colorNum, sizeNum, lineNum, dateNum = splits

    dayNum, monthNum = lot_number_functions.getDate(dateNum)
    
    global colorOut, boardOut
    colorOut = lot_number_functions.getColorDecode(colorNum)
    #print('')
    boardOut = lot_number_functions.getBoardSizeDecode(sizeNum)
    lineNum = int(lineNum, 2)

    lineNum = str(lineNum)
    #results = "Color: " + colorOut + "      Profile: " + boardOut + "\nDate: " + monthNum + '/' + dayNum + "        Line Number: " + lineNum + "\nPallet Num: " + palletEnd
    #result_label.config(text = results)
    lot_color_var.set(colorOut)
    lot_profile_var.set(boardOut)
    lot_pallet_num_var.set(palletEnd)
    lot_date_produced_var.set(monthNum + '/' + dayNum)
    lot_line_num_var.set(lineNum)
    color_lot_get()

def pull_from_file():
    print("pull_from_file")

# for recalling information from an existing lot number
def recall():
    try:  
        # Calls the function named submit_lot_num (not shown)
        submit_lot_num()
        set_recall_enabled(True)
        # Declares some global variables that will be used in the function
        global file_exists, file_path_pull, column_data
        
        # Sets the initial value of file_exists to False
        file_exists = False
        
        # Calls the read_file function from the recall_func module with arguments passed to it, 
        # and assigns the returned values to file_exists, column_data, time_keys, and file_path_pull
        file_exists, column_data, time_keys, file_path_pull = recall_func.read_file(lot_num.get(), colorOut, boardOut, lot_line_num_var.get())
        
        # Updates the values of the dropdown list with the new time keys
        hours_dropdown['values'] = time_keys

        # Sets the default value of the dropdown list to the first value in the list
        hours.set(time_keys[0])
        
        # Prints the value of file_exists to the console
        print(file_exists)
        
        # Prints the string "recalled" to the console
        print("recalled")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred, there likely is not a previous entry to recall")

#takes data from recall() and updatres tkinter 
def breakdown_data():
    # Prints the currently selected value from the hours_dropdown list to the console
    print(hours_dropdown.get())
    
    # Retrieves the data associated with the currently selected time value from the column_data dictionary
    data = column_data[hours_dropdown.get()]
    
    # Unpacks the retrieved data into individual variables for easier use
    [hour_sampled, date_sampled, o1, lot_num, profile, color, line_num, date_produced, pallet_num, o2, density, width_entry_recall, widthStr_recall, edge_entry_recall, edgeStr_recall,
                    middle_entry_recall, middleStr_recall, med_entry_recall, medStr_recall, o3, avg_L, range_L, avg_a, range_a, avg_b, range_b, deltaStr_recall, color_lot, notes_recall, o4, img_file_name_recall, ERC_file_name_recall] = data
    
    # Updates various tkinter variables with the retrieved data, making them display the correct values in the GUI
    densityStr.set(density)
    dateStr.set(date_sampled)
    width_entry_str.set(width_entry_recall)
    edge_entry_str.set(edge_entry_recall)
    middle_entry_str.set(middle_entry_recall)
    med_entry_str.set(med_entry_recall)
    deltaStr.set(deltaStr_recall)
    color_lot_number_entry["values"] = []
    color_lot_number_entry.set(color_lot)
    lot_avg_L.set(avg_L)
    lot_avg_a.set(avg_a)
    lot_avg_b.set(avg_b)
    lot_range_L.set(range_L)
    lot_range_a.set(range_a)
    lot_range_b.set(range_b)
    widthStr.set(widthStr_recall)
    edgeStr.set(edgeStr_recall)
    middleStr.set(middleStr_recall)
    medStr.set(medStr_recall)
    notes_entry_str.set(notes_recall)
    ERC_file_name.set(ERC_file_name_recall)
    img_file_name.set(img_file_name_recall)
    
    # Calls the on_index_click function (not shown)
    on_index_click()
    
    # Prints the retrieved data to the console
    print(data)

#when combobox is selected, check if in recall mode by looking for file exists, otherewise ignore
def is_recall(event):
    try:
        # code that uses the undefined variable
        if file_exists == True:
            breakdown_data()
        else:
            print(file_exists)
    except NameError:
        #ignore the error and continue with the rest of the code
        pass

def restart_program():
    python = sys.executable
    os.chdir(os.path.dirname(sys.executable))
    subprocess.Popen([python] + sys.argv)
    root.destroy()

submit_button = tk.Button(root, text="Submit", command=submit_lot_num)
submit_button.grid(row=1, column=2)
submit_button.config(font=("Helvetica", 12))

recall_button = tk.Button(root, text="Recall", command=recall)
recall_button.grid(row=1, column=3)
recall_button.config(font=("Helvetica", 12))

restart_button = tk.Button(root, text="Restart", command=restart_program)
restart_button.grid(row=1, column=5)
restart_button.config(font=("Helvetica", 10))

#output entered lot number for confirmation
lot_num_conf = tk.Label(root)
lot_num_conf.grid(row=3, column=1)
lot_num_conf.config(font=("Helvetica", 12))

#Color label and dynamic output from lot number
# lot Color label
color_static_label = tk.Label(root, text="Color:")
color_static_label.grid(row=4, column=0)
color_static_label.config(font=("Helvetica", 12))

# lot color variable to display
lot_color_var = tk.StringVar()
#lot_color_var.set("Variable Output")


# label that displays the lot color
lot_color_dynamic_label = tk.Label(root, textvariable=lot_color_var)
lot_color_dynamic_label.grid(row=4, column=1)
lot_color_dynamic_label.config(font=("Helvetica", 12))

#profile label and dynamic output from lot number
# lot profile label
profile_static_label = tk.Label(root, text="Profile:")
profile_static_label.grid(row=4, column=2)
profile_static_label.config(font=("Helvetica", 12))

# lot profile variable to display
lot_profile_var = tk.StringVar()
#lot_profile_var.set("Variable Output")


# label that displays the lot profile
lot_profile_dynamic_label = tk.Label(root, textvariable=lot_profile_var)
lot_profile_dynamic_label.grid(row=4, column=3)
lot_profile_dynamic_label.config(font=("Helvetica", 12))

#pallet_num label and dynamic output from lot number
# lot pallet_num label
pallet_num_static_label = tk.Label(root, text="Pallet #:")
pallet_num_static_label.grid(row=4, column=4)
pallet_num_static_label.config(font=("Helvetica", 12))

# lot pallet_num variable to display
lot_pallet_num_var = tk.StringVar()
#lot_pallet_num_var.set("Variable Output")

# label that displays the lot pallet_num
lot_pallet_num_dynamic_label = tk.Label(root, textvariable=lot_pallet_num_var)
lot_pallet_num_dynamic_label.grid(row=4, column=5)
lot_pallet_num_dynamic_label.config(font=("Helvetica", 12))

#date_produced label and dynamic output from lot number
# lot date_produced label
date_produced_static_label = tk.Label(root, text="Date Produced:")
date_produced_static_label.grid(row=5, column=0)
date_produced_static_label.config(font=("Helvetica", 12))

# lot date_produced variable to display
lot_date_produced_var = tk.StringVar()
#lot_date_produced_var.set("Variable Output")

# label that displays the lot date_produced
lot_date_produced_dynamic_label = tk.Label(root, textvariable=lot_date_produced_var)
lot_date_produced_dynamic_label.grid(row=5, column=1)
lot_date_produced_dynamic_label.config(font=("Helvetica", 12))

#line_num label and dynamic output from lot number
# lot line_num label
line_num_static_label = tk.Label(root, text="Line #:")
line_num_static_label.grid(row=5, column=2)
line_num_static_label.config(font=("Helvetica", 12))

# lot line_num variable to display
lot_line_num_var = tk.StringVar()
#lot_line_num_var.set("Variable Output")

# label that displays the lot line_num
lot_line_num_dynamic_label = tk.Label(root, textvariable=lot_line_num_var)
lot_line_num_dynamic_label.grid(row=5, column=3)
lot_line_num_dynamic_label.config(font=("Helvetica", 12))


#time sampled label
#time_sampled label and dynamic output from lot number
# lot time_sampled label
time_sampled_static_label = tk.Label(root, text="Time Sampled:")
time_sampled_static_label.grid(row=7, column=0)
time_sampled_static_label.config(font=("Helvetica", 12))

#for time sampled selection
def show_selected_value(*args):
    time_sampled = hours.get()
    #print(f"Selected hour: {time_sampled}")


hours = tk.StringVar()
hours.set("0700")
hours.trace("w", show_selected_value)

#combo dropdown box to select hour
hours_tuple = ("0000", "0100", "0200", "0300", "0400", "0500", "0600", "0700", "0800", "0900", 
                            "1000", "1100", "1200", "1300", "1400", "1500", "1600", "1700", "1800", "1900", 
                            "2000", "2100", "2200", "2300")
hours_dropdown = ttk.Combobox(root, textvariable=hours, values = hours_tuple)
hours_dropdown.bind("<<ComboboxSelected>>", is_recall)
hours_dropdown.current(7)
hours_dropdown.grid(row=7, column=1)
hours_dropdown.config(font=("Helvetica", 12))

# Call the function when the selection changes
#hours_dropdown.bind("<<ComboboxSelected>>", my_function)

#calandar
#calandar label
# density label with text entry next to it
calandar_label = tk.Label(root, text="Date:")
calandar_label.grid(row=7, column=3)
calandar_label.config(font=("Helvetica", 12))
# Declare string variable for date
dateStr = tk.StringVar()
date_picker = DateEntry(root, date_pattern="yyyy-MM-dd",textvariable=dateStr)
date_picker.set_date(datetime.date.today())
date_picker.config(font=("Helvetica", 12))
date_picker.grid(row=7, column=4)

# density label with text entry next to it
density_label = tk.Label(root, text="Density:")
density_label.grid(row=8, column=0)
density_label.config(font=("Helvetica", 12))

densityStr = tk.StringVar(value='0')
density_entry = tk.Entry(root, textvariable=densityStr)
density_entry.grid(row=8, column=1)
density_entry.config(font=("Helvetica", 12))

#width label and dynamic output from lot number
# lot width label
width_static_label = tk.Label(root, text="Width:")
width_static_label.grid(row=9, column=1)
width_static_label.config(font=("Helvetica", 12))

#edge label and dynamic output from lot number
# lot edge label
edge_static_label = tk.Label(root, text="Edge:")
edge_static_label.grid(row=9, column=2)
edge_static_label.config(font=("Helvetica", 12))

#middle label and dynamic output from lot number
# lot middle label
middle_static_label = tk.Label(root, text="Middle:")
middle_static_label.grid(row=9, column=3)
middle_static_label.config(font=("Helvetica", 12))

#med label and dynamic output from lot number
# lot med label
med_static_label = tk.Label(root, text="Med:")
med_static_label.grid(row=9, column=4)
med_static_label.config(font=("Helvetica", 12))

#delta label and dynamic output from lot number
# lot delta label
delta_static_label = tk.Label(root, text="Δ:")
delta_static_label.grid(row=9, column=5)
delta_static_label.config(font=("Helvetica", 12))


#board_profile label and dynamic output from lot number
# lot board_profile label
board_profile_static_label = tk.Label(root, text="Board Profile:")
board_profile_static_label.grid(row=10, column=0)
board_profile_static_label.config(font=("Helvetica", 12))

#text entry
width_entry_str = tk.StringVar(value='0')
width_entry = tk.Entry(root, textvariable=width_entry_str)
width_entry.grid(row=10, column=1)
width_static_label.config(font=("Helvetica", 12))

#text entry
edge_entry_str = tk.StringVar(value='0')
edge_entry = tk.Entry(root, textvariable=edge_entry_str)
edge_entry.grid(row=10, column=2)
edge_static_label.config(font=("Helvetica", 12))

#text entry
middle_entry_str = tk.StringVar(value='0')
middle_entry = tk.Entry(root, textvariable=middle_entry_str)
middle_entry.grid(row=10, column=3)
middle_static_label.config(font=("Helvetica", 12))

#text entry
med_entry_str = tk.StringVar(value='0')
med_entry = tk.Entry(root, textvariable=med_entry_str)
med_entry.grid(row=10, column=4)
med_static_label.config(font=("Helvetica", 12))

#results label and dynamic output from lot number
# lot results label
results_static_label = tk.Label(root, text="Results:")
results_static_label.grid(row=11, column=0)
results_static_label.config(font=("Helvetica", 12))

#output label with color change
widthStr = tk.StringVar()
width_variable_label = tk.Label(root, textvariable=widthStr)
width_variable_label.grid(row=11, column=1)
width_variable_label.config(font=("Helvetica", 12))

#output label with color change
edgeStr = tk.StringVar()
edge_variable_label = tk.Label(root, textvariable=edgeStr)
edge_variable_label.grid(row=11, column=2)
edge_variable_label.config(font=("Helvetica", 12))

#output label with color change
middleStr = tk.StringVar()
middle_variable_label = tk.Label(root, textvariable=middleStr)
middle_variable_label.grid(row=11, column=3)
middle_variable_label.config(font=("Helvetica", 12))

#output label with color change
medStr = tk.StringVar()
med_variable_label = tk.Label(root, textvariable=medStr)
med_variable_label.grid(row=11, column=4)
med_variable_label.config(font=("Helvetica", 12))

#output label with color change
deltaStr = tk.StringVar()
delta_variable_label = tk.Label(root, textvariable=deltaStr)
delta_variable_label.grid(row=11, column=5)
delta_variable_label.config(font=("Helvetica", 12))

#color_entry label and dynamic output from lot number
# lot color_entry label
color_entry_static_label = tk.Label(root, text="Color:")
color_entry_static_label.grid(row=13, column=0)
color_entry_static_label.config(font=("Helvetica", 12))

# color_lot_number label with text entry next to it
color_lot_number_label = tk.Label(root, text="Color Lot #:")
color_lot_number_label.grid(row=13, column=2)
color_lot_number_label.config(font=("Helvetica", 12))

# Create a combobox called "color_lot"
lot_color_dynamic_get = tk.StringVar()
color_lot_number_entry = ttk.Combobox(root, textvariable=lot_color_dynamic_get)
color_lot_number_entry.grid(row=13, column=3)
color_lot_number_entry.config(font=("Helvetica", 12))

#color_lotStr = tk.StringVar()
#color_lot_number_entry = tk.Entry(root, textvariable=color_lotStr)
#color_lot_number_entry.grid(row=13, column=3)
#color_lot_number_entry.config(font=("Helvetica", 12))

# L1 label with text entry next to it
L1_label = tk.Label(root, text="L1:")
L1_label.grid(row=14, column=0)
L1_label.config(font=("Helvetica", 12))


L1_entry = tk.Entry(root)
L1_entry.grid(row=14, column=1)
L1_entry.config(font=("Helvetica", 12))

# a1 label with text entry next to it
a1_label = tk.Label(root, text="a1:")
a1_label.grid(row=14, column=2)
a1_label.config(font=("Helvetica", 12))

a1_entry = tk.Entry(root)
a1_entry.grid(row=14, column=3)
a1_entry.config(font=("Helvetica", 12))

# b1 label with text entry next to it
b1_label = tk.Label(root, text="b1:")
b1_label.grid(row=14, column=4)
b1_label.config(font=("Helvetica", 12))

b1_entry = tk.Entry(root)
b1_entry.grid(row=14, column=5)
b1_entry.config(font=("Helvetica", 12))

# L2 label with text entry next to it
L2_label = tk.Label(root, text="L2:")
L2_label.grid(row=15, column=0)
L2_label.config(font=("Helvetica", 12))

L2_entry = tk.Entry(root)
L2_entry.grid(row=15, column=1)
L2_entry.config(font=("Helvetica", 12))

# a2 label with text entry next to it
a2_label = tk.Label(root, text="a2:")
a2_label.grid(row=15, column=2)
a2_label.config(font=("Helvetica", 12))

a2_entry = tk.Entry(root)
a2_entry.grid(row=15, column=3)
a2_entry.config(font=("Helvetica", 12))

# b2 label with text entry next to it
b2_label = tk.Label(root, text="b2:")
b2_label.grid(row=15, column=4)
b2_label.config(font=("Helvetica", 12))

b2_entry = tk.Entry(root)
b2_entry.grid(row=15, column=5)
b2_entry.config(font=("Helvetica", 12))

# L3 label with text entry next to it
L3_label = tk.Label(root, text="L3:")
L3_label.grid(row=16, column=0)
L3_label.config(font=("Helvetica", 12))

L3_entry = tk.Entry(root)
L3_entry.grid(row=16, column=1)
L3_entry.config(font=("Helvetica", 12))

# a3 label with text entry next to it
a3_label = tk.Label(root, text="a3:")
a3_label.grid(row=16, column=2)
a3_label.config(font=("Helvetica", 12))

a3_entry = tk.Entry(root)
a3_entry.grid(row=16, column=3)
a3_entry.config(font=("Helvetica", 12))

# b3 label with text entry next to it
b3_label = tk.Label(root, text="b3:")
b3_label.grid(row=16, column=4)
b3_label.config(font=("Helvetica", 12))

b3_entry = tk.Entry(root)
b3_entry.grid(row=16, column=5)
b3_entry.config(font=("Helvetica", 12))

#avg_L label and dynamic output from lot number
# lot avg_L label
avg_L_static_label = tk.Label(root, text="AVG L:")
avg_L_static_label.grid(row=18, column=0)
avg_L_static_label.config(font=("Helvetica", 12))


lot_avg_L = tk.StringVar(value='0')
# lot avg_L variable to display
lot_avg_L_var = tk.Entry(root, textvariable=lot_avg_L)
lot_avg_L_var.grid(row=18, column=1)
lot_avg_L_var.config(font=("Helvetica", 12))

lot_range_L = tk.StringVar()
# label that displays the lot avg_L
lot_avg_L_dynamic_label = tk.Label(root, textvariable=lot_range_L)
lot_avg_L_dynamic_label.grid(row=19, column=1)
lot_avg_L_dynamic_label.config(font=("Helvetica", 12))

#avg_a label and dynamic output from lot number
# lot avg_a label
avg_a_static_label = tk.Label(root, text="AVG a:")
avg_a_static_label.grid(row=18, column=2)
avg_a_static_label.config(font=("Helvetica", 12))

lot_avg_a = tk.StringVar(value='0')
# lot avg_a variable to display
lot_avg_a_var = tk.Entry(root, textvariable=lot_avg_a)
lot_avg_a_var.grid(row=18, column=3)
lot_avg_a_var.config(font=("Helvetica", 12))

lot_range_a = tk.StringVar()
# label that displays the lot avg_a
lot_avg_a_dynamic_label = tk.Label(root, textvariable=lot_range_a)
lot_avg_a_dynamic_label.grid(row=19, column=3)
lot_avg_a_dynamic_label.config(font=("Helvetica", 12))

#avg_b label and dynamic output from lot number
# lot avg_b label
avg_b_static_label = tk.Label(root, text="AVG b:")
avg_b_static_label.grid(row=18, column=4)
avg_b_static_label.config(font=("Helvetica", 12))

lot_avg_b = tk.StringVar(value='0')
# lot avg_b variable to display
lot_avg_b_var = tk.Entry(root, textvariable=lot_avg_b)
lot_avg_b_var.grid(row=18, column=5)
lot_avg_b_var.config(font=("Helvetica", 12))

lot_range_b = tk.StringVar()
# label that displays the lot avg_b
lot_avg_b_dynamic_label = tk.Label(root, textvariable=lot_range_b)
lot_avg_b_dynamic_label.grid(row=19, column=5)
lot_avg_b_dynamic_label.config(font=("Helvetica", 12))

# Notes label with text entry next to it

notes_label = tk.Label(root, text="Notes:")
notes_label.grid(row=20, column=1)
notes_label.config(font=("Helvetica", 12))

notes_entry_str = tk.StringVar()
notes_entry = tk.Entry(root, textvariable=notes_entry_str)
notes_entry.grid(row=20, column=2)
notes_entry.config(font=("Helvetica", 12))

#select extruder running conditons file
def select_ERC_file():
    global pdf_file_path
    pdf_file_path = filedialog.askopenfilename(initialdir = "/Desktop", defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
    if pdf_file_path:
        ERC_file_name.set(pdf_file_path)

ERC_file_name = tk.StringVar()

select_button = tk.Button(root, text="Select ERC PDF", command=select_ERC_file)
select_button.grid(row=21, column=3)
select_button.config(font=("Helvetica", 12))


file_label = tk.Label(root, textvariable=ERC_file_name)
file_label.grid(row=22, column=3)
file_label.config(font=("Helvetica", 12))

#select pallet photo

def select_image_file():
        
    global file_path
    lot_num = lot_num_entry.get()
    file_path = filedialog.askopenfilename(initialdir = "/Desktop", defaultextension = ".jpg", filetypes=[("Image Files", "*.jpg;*.jpeg;*.png;*.gif")])
    # Define the new file name and location
    new_file_name = lot_num + '_photo.jpg'
    new_file_folder = Data_Locations.set_photo_place()
    new_file_path = new_file_folder + new_file_name
    # Copy the image to the new location with the new file name
    shutil.copy(file_path, new_file_path)
    file_path = new_file_path
    if file_path:
        img_file_name.set(file_path)

img_file_name = tk.StringVar()

photo_select_button = tk.Button(root, text="Select Lot Image", command=select_image_file)
photo_select_button.grid(row=21, column=1)
photo_select_button.config(font=("Helvetica", 12))

photo_file_label = tk.Label(root, textvariable=img_file_name)
photo_file_label.grid(row=22, column=1)
photo_file_label.config(font=("Helvetica", 12))

def open_image():
    file_path = img_file_name.get()
    if file_path:
        try:
            image = Image.open(file_path)
            image.show()
        except Exception as e:
            messagebox.showerror("Error", "Failed to open the image\n" + str(e))


image_hyperlink = ttk.Button(root, text="Show Image", command=open_image)
image_hyperlink.grid(row=24, column=1)


def open_pdf():
    file_path = ERC_file_name.get()
    print(file_path)
    subprocess.Popen(["rundll32", "url.dll,FileProtocolHandler", file_path.replace("/", "\\")])
    

pdf_hyperlink = ttk.Button(root, text="Open PDF", command=open_pdf)
pdf_hyperlink.grid(row=24, column=3)

#button to get and store variables
def on_index_click():
    # This function is called when the "Store" button is clicked
    global avgL, avga, avgb, L_range, a_range, b_range
    print("Index button clicked")
    
    # Get variables from user inputs
    profile_width = width_entry.get()
    profile_edge = edge_entry.get()
    profile_middle = middle_entry.get()
    profile_med = med_entry.get()

    # Check profile standard range
    result_thickness_T, result_thickness_B, result_width_T, result_width_B = Data_Locations.check_profile_standard_range(profile_width, profile_edge, profile_middle, profile_med, boardOut)
    pass_profile_tuple = (result_thickness_T, result_thickness_B, result_width_T, result_width_B, profile_width, profile_edge, profile_middle, profile_med)
    profile_tuple = Data_Locations.profile_range_check(pass_profile_tuple)
    width_range, pass_range_width, edge_range, pass_range_edge, middle_range, pass_range_middle, med_range, pass_range_med, delta_range, pass_range_delta = profile_tuple

    # Round values to 2 decimal places
    width_range = round(width_range, 2)
    edge_range = round(edge_range, 2)
    middle_range = round(middle_range, 2)
    med_range = round(med_range, 2)
    delta_range = round(delta_range, 2)

    # Set text and color for widthStr
    if pass_range_width:
        widthStr.set("Pass")
        width_variable_label.config(font=("Helvetica", 12), fg="green")
    else:
        widthStr.set(width_range)
        width_variable_label.config(font=("Helvetica", 12), fg="red")

    # Set text and color for edgeStr
    if pass_range_edge:
        edgeStr.set("Pass")
        edge_variable_label.config(font=("Helvetica", 12), fg="green")
    else:
        edgeStr.set(edge_range)
        edge_variable_label.config(font=("Helvetica", 12), fg="red")

    # Set text and color for middleStr
    if pass_range_middle:
        middleStr.set("Pass")
        middle_variable_label.config(font=("Helvetica", 12), fg="green")
    else:
        middleStr.set(middle_range)
        middle_variable_label.config(font=("Helvetica", 12), fg="red")

    # Set text and color for medStr
    if pass_range_med:
        medStr.set("Pass")
        med_variable_label.config(font=("Helvetica", 12), fg="green")
    else:
        medStr.set(med_range)
        med_variable_label.config(font=("Helvetica", 12), fg="red")

    # Set text and color for deltaStr
    if pass_range_delta == True:
        deltaStr.set("Pass")
        delta_variable_label.config(font=("Helvetica", 12), fg="green")
    else:
        deltaStr.set(delta_range)
        delta_variable_label.config(font=("Helvetica", 12), fg="red")

    ######Color Check######
    color_L1 = L1_entry.get()
    color_L2 = L2_entry.get()
    color_L3 = L3_entry.get()
    color_a1 = a1_entry.get()
    color_a2 = a2_entry.get()
    color_a3 = a3_entry.get()
    color_b1 = b1_entry.get()
    color_b2 = b2_entry.get()
    color_b3 = b3_entry.get()
    avg_L_set = lot_avg_L.get()
    avg_a_set = lot_avg_a.get()
    avg_b_set = lot_avg_b.get()
    
    # Combine the color variables into a tuple
    if color_L1 != "":
        colorTuple = (color_L1, color_L2, color_L3, color_a1, color_a2, color_a3, color_b1, color_b2, color_b3)
    elif avg_L_set != "":
        colorTuple = (avg_L_set, avg_a_set, avg_b_set)
    else:
        print("color entry error")
    
    # Calculate the average L, a, and b values
    avgL, avga, avgb = Data_Locations.averageLab(colorTuple)
    
    # Round the average values to 2 decimal places
    avgL = round(avgL, 2)
    avga = round(avga, 2)
    avgb = round(avgb, 2)

    # Update the UI with the average values
    lot_avg_L.set(avgL)
    lot_avg_a.set(avga)
    lot_avg_b.set(avgb)

    # Check if the average values are within the color standard range
    result_L_top, result_a_top, result_b_top, result_L_bottom, result_a_bottom, result_b_bottom = Data_Locations.check_color_standard_range(avgL, avga, avgb, colorOut)

    # Combine the result values into a tuple
    color_result_tuple = (result_L_top, result_a_top, result_b_top, result_L_bottom, result_a_bottom, result_b_bottom, avgL, avga, avgb)
    
    pass_range_tuple = Data_Locations.color_range_check(color_result_tuple)

    L_range, a_range, b_range, pass_rangeL, pass_rangea, pass_rangeb = pass_range_tuple

        # Round the result values to 2 decimal places
    L_range = round(float(L_range), 2)
    a_range = round(float(a_range), 2)
    b_range = round(float(b_range), 2)

    if pass_rangeL == True:
        lot_range_L.set("Pass")
        lot_avg_L_dynamic_label.config(font=("Helvetica", 12), fg="green")
    else:
        lot_range_L.set(L_range)
        lot_avg_L_dynamic_label.config(font=("Helvetica", 12), fg="red")

    if pass_rangea == True:
        lot_range_a.set("Pass")
        lot_avg_a_dynamic_label.config(font=("Helvetica", 12), fg="green")
    else:
        lot_range_a.set(a_range)
        lot_avg_a_dynamic_label.config(font=("Helvetica", 12), fg="red")

    if pass_rangeb == True:
        lot_range_b.set("Pass")
        lot_avg_b_dynamic_label.config(font=("Helvetica", 12), fg="green")
    else:
        lot_range_b.set(b_range)
        lot_avg_b_dynamic_label.config(font=("Helvetica", 12), fg="red")

######################################
#confirm window stuff
def continue_function():
    """Function to execute if user wants to continue"""
    messagebox.showinfo("Message", "Continuing...")

def continue_window():
    """Function to create a new window"""
    new_window = tk.Toplevel()
    new_window.title("Overwrite?")
    new_window.geometry("200x100")

    # Ask user if they want to continue
    response = messagebox.askquestion("Overwrite Existing Information?", "Do you want to continue?")
    if response == "yes":
        new_window.destroy()
        return True
    else:
        new_window.destroy()
        return False

###################################
#stores data to files when clicked
def on_store_click():
    on_index_click()
    print("Store button clicked")
    global recall_enabled
    print(recall_enabled)
    if recall_enabled == True:
        continue_question = continue_window()
        if continue_question == True:
            print('continuing')

            format_excel_tuple = (lot_num.get(), lot_color_var.get(), lot_profile_var.get(), lot_pallet_num_var.get(), lot_date_produced_var.get(), lot_line_num_var.get(), hours.get(), dateStr.get())
            to_excel_tuple = (lot_num.get(), lot_color_var.get(), lot_profile_var.get(), lot_pallet_num_var.get(), lot_date_produced_var.get(), lot_line_num_var.get(), hours.get(), dateStr.get(),
                                densityStr.get(), lot_avg_L.get(), lot_avg_a.get(), lot_avg_b.get(), lot_range_L.get(), lot_range_a.get(), lot_range_b.get(), deltaStr.get(),
                                    width_entry.get(), edge_entry.get(), middle_entry.get(), med_entry.get(), widthStr.get(), edgeStr.get(), middleStr.get(), medStr.get(), deltaStr.get(), 
                                        lot_color_dynamic_get.get(), notes_entry.get(), img_file_name.get(),  ERC_file_name.get())
            file_loc = Data_Locations.format_excel(format_excel_tuple)
            Data_Locations.write_to_excel(file_loc, to_excel_tuple)
            
            Data_Locations.write_color(to_excel_tuple)
            Data_Locations.write_profile(to_excel_tuple)

            # Create a new window to show "success" message
            success_window = tk.Toplevel()
            success_label = tk.Label(success_window, text="Success!")
            success_label.pack()

        elif continue_question == False:
            print('stopping')
    if recall_enabled == False:
            format_excel_tuple = (lot_num.get(), lot_color_var.get(), lot_profile_var.get(), lot_pallet_num_var.get(), lot_date_produced_var.get(), lot_line_num_var.get(), hours.get(), dateStr.get())
            to_excel_tuple = (lot_num.get(), lot_color_var.get(), lot_profile_var.get(), lot_pallet_num_var.get(), lot_date_produced_var.get(), lot_line_num_var.get(), hours.get(), dateStr.get(),
                                densityStr.get(), lot_avg_L.get(), lot_avg_a.get(), lot_avg_b.get(), lot_range_L.get(), lot_range_a.get(), lot_range_b.get(), deltaStr.get(),
                                    width_entry.get(), edge_entry.get(), middle_entry.get(), med_entry.get(), widthStr.get(), edgeStr.get(), middleStr.get(), medStr.get(), deltaStr.get(), 
                                        lot_color_dynamic_get.get(), notes_entry.get(), img_file_name.get(),  ERC_file_name.get())
            file_loc = Data_Locations.format_excel(format_excel_tuple)
            Data_Locations.write_to_excel(file_loc, to_excel_tuple)
            
            Data_Locations.write_color(to_excel_tuple)
            Data_Locations.write_profile(to_excel_tuple)

            # Create a new window to show "success" message
            success_window = tk.Toplevel()
            success_label = tk.Label(success_window, text="Success!")
            success_label.pack()

index_button = tk.Button(root, text="Index", command=on_index_click)
index_button.grid(row=22, column=5)
index_button.config(font=("Helvetica", 12))

store_button = tk.Button(root, text="Store", command=on_store_click)
store_button.grid(row=23, column=5)
store_button.config(font=("Helvetica", 12))

#add blank rows for spacing
blank_row0 = tk.Label(root, text="", height=2)
blank_row0.grid(row=3, column=0)

blank_row1 = tk.Label(root, text="", height=2)
blank_row1.grid(row=6, column=0)

blank_row2 = tk.Label(root, text="", height=2)
blank_row2.grid(row=12, column=0)

blank_row3 = tk.Label(root, text="", height=2)
blank_row3.grid(row=20, column=0)


root.mainloop()